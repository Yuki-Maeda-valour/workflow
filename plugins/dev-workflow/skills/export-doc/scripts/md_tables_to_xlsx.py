#!/usr/bin/env python3
"""Markdown 文書内の表を xlsx(または CSV)へ変換する(export-doc 用)。

使い方:
  python3 md_tables_to_xlsx.py <input.md> [<input2.md> ...] -o <output.xlsx>

- 1 表 1 シート。シート名は「文書名 + 直前の見出し」(Excel の 31 文字制約に収める)
- openpyxl が無い環境では、出力名と同名のディレクトリへ CSV(BOM 付き UTF-8)で縮退出力する
- 入力は読み取りのみ(変更しない)。同名出力は上書き(冪等)
- セル内の Markdown 装飾(**強調**・`コード`・[リンク](url)・<br>)はプレーンテキスト化する

終了コード: 0 = 出力あり / 1 = 表が見つからない / 2 = 入力エラー
"""

import argparse
import csv
import re
import sys
from pathlib import Path

CELL_CLEAN = [
    (re.compile(r"\*\*(.+?)\*\*"), r"\1"),
    (re.compile(r"`([^`]*)`"), r"\1"),
    (re.compile(r"\[([^\]]*)\]\([^)]*\)"), r"\1"),
    (re.compile(r"<br\s*/?>", re.IGNORECASE), "\n"),
]


def clean_cell(s: str) -> str:
    s = s.strip()
    for pat, rep in CELL_CLEAN:
        s = pat.sub(rep, s)
    return s


def split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    parts = re.split(r"(?<!\\)\|", line)
    return [clean_cell(p.replace("\\|", "|")) for p in parts]


def is_separator(line: str) -> bool:
    body = line.strip().strip("|")
    cells = [c.strip() for c in body.split("|")]
    filled = [c for c in cells if c]
    return bool(filled) and all(re.fullmatch(r":?-{3,}:?", c) for c in filled)


def extract_tables(text: str, doc_name: str) -> list[tuple[str, list[list[str]]]]:
    tables = []
    lines = text.splitlines()
    heading = ""
    in_code = False
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.lstrip().startswith("```"):
            in_code = not in_code
            i += 1
            continue
        if in_code:
            i += 1
            continue
        m = re.match(r"^(#{1,6})\s+(.*)", line)
        if m:
            heading = m.group(2).strip()
            i += 1
            continue
        if line.lstrip().startswith("|") and i + 1 < len(lines) and is_separator(lines[i + 1]):
            rows = [split_row(line)]
            i += 2
            while i < len(lines) and lines[i].lstrip().startswith("|"):
                rows.append(split_row(lines[i]))
                i += 1
            tables.append((f"{doc_name} {heading}".strip(), rows))
            continue
        i += 1
    return tables


def unique_sheet_name(name: str, used: set) -> str:
    name = re.sub(r"[\\/*?\[\]:]", " ", name).strip() or "Sheet"
    cand = name[:31]
    n = 2
    while cand in used:
        suffix = f"_{n}"
        cand = name[: 31 - len(suffix)] + suffix
        n += 1
    used.add(cand)
    return cand


def write_csv_fallback(tables, out: Path) -> None:
    outdir = out.with_suffix("")
    outdir.mkdir(parents=True, exist_ok=True)
    used = set()
    for hint, rows in tables:
        name = unique_sheet_name(hint, used)
        with (outdir / f"{name}.csv").open("w", newline="", encoding="utf-8-sig") as fh:
            csv.writer(fh).writerows(rows)
    print(f"FALLBACK openpyxl 未導入のため CSV で出力: {outdir}/ (pip install openpyxl で xlsx 化可)")


def write_xlsx(tables, out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for hint, rows in tables:
        ws = wb.create_sheet(unique_sheet_name(hint, used))
        for r in rows:
            ws.append(r)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            width = 8
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value or ""
                first_line = str(v).split("\n")[0]
                w = sum(2 if ord(ch) > 0x7F else 1 for ch in first_line)
                width = max(width, min(w + 2, 60))
            ws.column_dimensions[get_column_letter(col)].width = width
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"OK {out} にシート {len(wb.sheetnames)} 件を出力")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("inputs", nargs="+", help="入力 Markdown ファイル")
    ap.add_argument("-o", "--output", required=True, help="出力 xlsx パス")
    args = ap.parse_args()

    tables = []
    missing = []
    for p in args.inputs:
        path = Path(p)
        if not path.is_file():
            missing.append(p)
            continue
        tables += extract_tables(path.read_text(encoding="utf-8"), path.stem)
    for p in missing:
        print(f"WARN 入力が存在しない: {p}", file=sys.stderr)
    if missing and not tables:
        return 2
    if not tables:
        print("表が見つからなかった(出力なし)")
        return 1

    out = Path(args.output)
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        write_csv_fallback(tables, out)
        return 0
    write_xlsx(tables, out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
